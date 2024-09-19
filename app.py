import tkinter as tk
from tkinter import ttk, messagebox, IntVar, filedialog
import ttkbootstrap as ttkb
from ttkbootstrap.constants import *
import pandas as pd
import random
import os
import json
import matplotlib.pyplot as plt
from fpdf import FPDF
from tkinter import simpledialog
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl import Workbook

# Função para gerar gráficos comparativos das notas medianas dos frameworks
def gerar_grafico_comparativo_multiplos_avaliadores(frameworks):
    try:
        for framework in frameworks:
            df_criterios = pd.DataFrame()
            for avaliacao in framework['avaliacoes']:
                df = avaliacao['df_criterios'].copy()
                df['Avaliador'] = avaliacao['Avaliador']
                df_criterios = pd.concat([df_criterios, df], ignore_index=True)

            if not df_criterios.empty:
                # Gráfico comparando as notas dos avaliadores para cada critério por framework
                df_pivot = df_criterios.pivot_table(index='Critério', columns='Avaliador', values='Nota do Critério', aggfunc='mean')
                df_pivot.plot(kind='bar', figsize=(12, 7), colormap='Set3', edgecolor='black')
                plt.title(f'Comparação de Notas entre Avaliadores - {framework["Framework"]}')
                plt.ylabel('Nota (1-5)')
                plt.xticks(rotation=45, ha='right')
                plt.tight_layout()
                plt.show()

    except Exception as e:
        messagebox.showerror("Erro ao gerar gráfico", f"Ocorreu um erro ao gerar o gráfico: {e}")




# Classe principal da aplicação
class App(ttkb.Window):
    def __init__(self):
        super().__init__(themename="flatly")
        self.title("Sistema de Avaliação")
        self.geometry("800x600")
        self.configure(bg='#e0e0e0')

        # Inicialize as listas para armazenar os widgets Entry
        self.nomes_avaliadores_entries = []  # Armazena as entradas de nomes de avaliadores
        self.nomes_frameworks_entries = []  # Armazena as entradas de nomes de frameworks

        self.criterios = [
            "Custo", "Segurança da Informação", "Eficiência", "Desempenho", 
            "Complexidade", "Flexibilidade/Adaptabilidade", "Conformidade", 
            "Suporte e Documentação", "Escalabilidade", "Comunidade e Adoção", 
            "Integração com Outras Ferramentas", "Inovação e Atualização"
        ]

        self.subcriterios_default = {
            "Custo": ["Implementação", "Licença", "Treinamento", "Manutenção", "Consultoria"],
            "Segurança da Informação": ["Proteção de Dados", "Detecção de Intrusões", "Resposta a Incidentes", "Recuperação", "Prevenção"],
            "Eficiência": ["Otimização de Recursos", "Tempo de Resposta", "Automatização", "Escalabilidade", "Integração"],
            "Desempenho": ["Efetividade das Medidas de Segurança", "Taxa de Detecção de Ameaças", "Mitigação de Riscos", "Impacto na Operação", "Tempo de Recuperação"],
            "Complexidade": ["Facilidade de Implementação", "Curva de Aprendizado", "Requisitos Técnicos", "Compatibilidade com Sistemas Existentes", "Complexidade de Manutenção"],
            "Flexibilidade/Adaptabilidade": ["Adaptação a Diferentes Setores", "Customização", "Escalabilidade", "Integração com Outras Ferramentas", "Ajustes de Configuração"],
            "Conformidade": ["Regulamentação", "Políticas Internas", "Auditoria", "Relatórios", "Certificação"],
            "Suporte e Documentação": ["Qualidade da Documentação", "Disponibilidade de Suporte Técnico", "Comunidade de Usuários", "Recursos de Aprendizado", "Atualizações de Documentação"],
            "Escalabilidade": ["Capacidade de Crescimento", "Performance em Larga Escala", "Flexibilidade de Expansão", "Gerenciamento de Crescimento", "Suporte a Multinacionais"],
            "Comunidade e Adoção": ["Popularidade", "Feedback da Comunidade", "Exemplos de Uso Real", "Colaborações e Parcerias", "Desenvolvimento Contínuo"],
            "Integração com Outras Ferramentas": ["Compatibilidade", "APIs e Conectores", "Interoperabilidade", "Facilidade de Integração", "Suporte a Padrões Abertos"],
            "Inovação e Atualização": ["Frequência de Atualizações", "Incorporação de Novas Tecnologias", "Pesquisa e Desenvolvimento", "Feedback do Mercado", "Melhorias Contínuas"]
        }

         # Começa com a tela de identificação
        self.criar_tela_identificacao()

    # Função para criar a tela de nomes dos avaliadores e frameworks
    def criar_tela_identificacao(self):
        self.identificacao_frame = ttkb.Frame(self)  # Certifique-se de que o frame é salvo como self.identificacao_frame
        self.identificacao_frame.pack(fill='both', expand=True, padx=20, pady=20)

        # Perguntar quantos avaliadores e frameworks
        ttkb.Label(self.identificacao_frame, text="Quantos avaliadores?", font=("Helvetica", 12)).pack(pady=10)
        self.num_avaliadores = ttkb.Entry(self.identificacao_frame, width=10)
        self.num_avaliadores.pack(pady=5)

        ttkb.Label(self.identificacao_frame, text="Quantos frameworks?", font=("Helvetica", 12)).pack(pady=10)
        self.num_frameworks = ttkb.Entry(self.identificacao_frame, width=10)
        self.num_frameworks.pack(pady=5)

        # Escolha do modo de avaliação: manual ou simulação
        mode_frame = ttkb.Frame(self.identificacao_frame)
        mode_frame.pack(pady=20)
        ttkb.Label(mode_frame, text="Modo de Avaliação:", font=("Helvetica", 14, "bold")).grid(row=0, column=0, columnspan=2, pady=10)

        self.modo_avaliacao_var = tk.StringVar(value="manual")
        manual_button = ttkb.Radiobutton(mode_frame, text="Manual", variable=self.modo_avaliacao_var, value="manual", bootstyle="primary", width=10)
        manual_button.grid(row=1, column=0, padx=10, pady=5)
        simulacao_button = ttkb.Radiobutton(mode_frame, text="Simulação", variable=self.modo_avaliacao_var, value="simulacao", bootstyle="danger", width=10)
        simulacao_button.grid(row=1, column=1, padx=10, pady=5)

        # Botão para continuar
        ttkb.Button(self.identificacao_frame, text="Continuar", bootstyle="success", command=self.proxima_tela).pack(pady=20)

    def proxima_tela(self):
        try:
            # Verifica os valores de número de avaliadores e frameworks
            self.num_avaliadores_val = int(self.num_avaliadores.get().strip())
            self.num_frameworks_val = int(self.num_frameworks.get().strip())

            if self.num_avaliadores_val <= 0 or self.num_frameworks_val <= 0:
                raise ValueError("O número de avaliadores e frameworks deve ser maior que zero.")

            # Destrói a tela de identificação
            self.identificacao_frame.destroy()

            # Cria a tela de nomes para avaliadores e frameworks
            self.criar_tela_nomes(self.num_avaliadores_val, self.num_frameworks_val)

        except ValueError:
            messagebox.showerror("Erro", "Por favor, insira valores válidos.")
        except Exception as e:
            messagebox.showerror("Erro inesperado", f"Erro inesperado: {e}")




    def criar_tela_nomes(self, num_avaliadores, num_frameworks):
        # Destruir qualquer frame existente
        if hasattr(self, 'nomes_frame'):
            self.nomes_frame.destroy()

        self.nomes_frame = ttkb.Frame(self)
        self.nomes_frame.pack(fill='both', expand=True, padx=20, pady=20)

        # Scrollbar
        canvas = tk.Canvas(self.nomes_frame)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        scrollbar = ttkb.Scrollbar(self.nomes_frame, orient=tk.VERTICAL, command=canvas.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        scroll_frame = ttkb.Frame(canvas)
        canvas.create_window((0, 0), window=scroll_frame, anchor='nw')

        scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        self.nomes_avaliadores_entries = []
        self.nomes_frameworks_entries = []

        for i in range(num_avaliadores):
            ttkb.Label(scroll_frame, text=f"Nome do Avaliador {i + 1}:", font=("Helvetica", 12)).pack(pady=5)
            entry_avaliador = ttkb.Entry(scroll_frame, width=40)
            entry_avaliador.pack(pady=5)
            self.nomes_avaliadores_entries.append(entry_avaliador)

        for i in range(num_frameworks):
            ttkb.Label(scroll_frame, text=f"Nome do Framework {i + 1}:", font=("Helvetica", 12)).pack(pady=5)
            entry_framework = ttkb.Entry(scroll_frame, width=40)
            entry_framework.pack(pady=5)
            self.nomes_frameworks_entries.append(entry_framework)

        ttkb.Button(scroll_frame, text="Continuar", bootstyle="success", command=lambda: self.proxima_tela_criterios(num_frameworks)).pack(pady=20)


    # Definir o método proxima_tela_criterios, que estava ausente
    def proxima_tela_criterios(self, num_frameworks):
        try:
            # Certifique-se de que todos os nomes de avaliadores e frameworks estão preenchidos
            self.nomes_avaliadores = [entry.get() for entry in self.nomes_avaliadores_entries]
            self.nomes_frameworks = [entry.get() for entry in self.nomes_frameworks_entries]

            if len(self.nomes_avaliadores) < self.num_avaliadores_val or len(self.nomes_frameworks) < self.num_frameworks_val:
                raise ValueError("Por favor, preencha o nome de todos os avaliadores e frameworks.")

            # Destruir a tela de nomes (anterior)
            self.nomes_frame.destroy()

            # Iniciar a tela de avaliação
            self.iniciar_avaliacao(0)

        except ValueError as e:
            messagebox.showerror("Erro", str(e))
        except Exception as e:
            messagebox.showerror("Erro inesperado", str(e))


    def iniciar_avaliacao(self, avaliador_idx=0):
        if avaliador_idx < len(self.nomes_avaliadores):
            avaliador = self.nomes_avaliadores[avaliador_idx]
            messagebox.showinfo("Próximo Avaliador", f"Avaliador {avaliador}, por favor, inicie sua avaliação.")
            self.avaliador_atual = avaliador_idx

            # Destruir qualquer tela anterior
            if hasattr(self, 'criterios_frame'):
                self.criterios_frame.destroy()

            # Iniciar a tela de critérios
            self.criar_tela_criterios(self.criterios, self.subcriterios_default, self.modo_avaliacao_var.get(), len(self.nomes_avaliadores), self.num_frameworks_val)
        else:
            messagebox.showinfo("Avaliação Concluída", "Todos os avaliadores completaram a avaliação.")
            self.calcular_notas_multiplos_avaliadores()


    # Defina o método criar_tela_criterios na classe
    def criar_tela_criterios(self, criterios, subcriterios_default, modo_avaliacao, num_avaliadores, num_frameworks):
        # Destruir a tela anterior, se existir
        if hasattr(self, 'criterios_frame'):
            self.criterios_frame.destroy()

        # Criar novo frame para critérios
        self.criterios_frame = ttkb.Frame(self)
        self.criterios_frame.pack(fill='both', expand=True, padx=20, pady=20)

        # Adicionar barra de rolagem
        canvas = tk.Canvas(self.criterios_frame)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        scrollbar = ttkb.Scrollbar(self.criterios_frame, orient=tk.VERTICAL, command=canvas.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        scroll_frame = ttkb.Frame(canvas)
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        notebook = ttkb.Notebook(scroll_frame, bootstyle="info")
        notebook.pack(fill='both', expand=True, pady=10)

        self.pesos_criterios = {criterio: [] for criterio in criterios}
        self.pesos_subcriterios = {criterio: {} for criterio in criterios}
        self.pontuacoes_subcriterios = {criterio: {} for criterio in criterios}

        # Iterar sobre todos os avaliadores
        for avaliador_idx in range(num_avaliadores):
            # Criar um frame para cada avaliador
            frame_avaliador = ttkb.Frame(notebook)
            notebook.add(frame_avaliador, text=f"Avaliador {avaliador_idx + 1}")

            # Iterar sobre todos os frameworks
            for fw_idx in range(num_frameworks):  # Corrigido: garantir que loop seja feito para cada framework
                frame = ttkb.LabelFrame(frame_avaliador, text=f"Framework {fw_idx + 1}", padding=10)  # Certificar que cria frames para todos os frameworks
                frame.pack(fill='x', expand=True, padx=10, pady=10)

                # Criação das entradas de critérios e subcritérios para cada framework
                for criterio in criterios:
                    criterio_frame = ttkb.LabelFrame(frame, text=criterio, padding=10)
                    criterio_frame.pack(fill='x', expand=True, padx=10, pady=10)

                    ttkb.Label(criterio_frame, text="Peso do Critério:", font=("Helvetica", 10)).grid(row=0, column=0, sticky='w')
                    entry_peso = ttkb.Entry(criterio_frame, width=5)
                    entry_peso.grid(row=0, column=1, padx=10)
                    self.pesos_criterios[criterio].append(entry_peso)

                    ttkb.Label(criterio_frame, text="Subcritérios", font=("Helvetica", 10, "bold")).grid(row=1, column=0, columnspan=6, pady=5)

                    subcriterios = subcriterios_default.get(criterio, [])
                    sub_pesos = {}
                    sub_pontuacoes = {}
                    for i, subcriterio in enumerate(subcriterios):
                        ttkb.Label(criterio_frame, text=subcriterio).grid(row=i + 2, column=0, sticky='w', padx=10)

                        # Entrada de pontuação para subcritério
                        sub_pontuacao_var = IntVar(value=3)
                        for j in range(1, 6):
                            radiobutton = ttkb.Radiobutton(criterio_frame, text=str(j), variable=sub_pontuacao_var, value=j, bootstyle="info")
                            radiobutton.grid(row=i + 2, column=j, padx=5)

                        if subcriterio not in self.pontuacoes_subcriterios[criterio]:
                            self.pontuacoes_subcriterios[criterio][subcriterio] = []
                        self.pontuacoes_subcriterios[criterio][subcriterio].append(sub_pontuacao_var)

                        sub_peso = ttkb.Entry(criterio_frame, width=5)
                        sub_peso.grid(row=i + 2, column=6, padx=10)

                        if subcriterio not in self.pesos_subcriterios[criterio]:
                            self.pesos_subcriterios[criterio][subcriterio] = []
                        self.pesos_subcriterios[criterio][subcriterio].append(sub_peso)

        ttkb.Button(scroll_frame, text="Calcular Nota Final", bootstyle="success", command=self.calcular_notas).pack(pady=20)
        ttkb.Button(scroll_frame, text="Voltar", bootstyle="warning", command=self.voltar_para_identificacao).pack(pady=10)

        if modo_avaliacao == "simulacao":
            self.simular_valores_criterios_subcriterios(criterios, subcriterios_default, num_frameworks, num_avaliadores)

    # Função para simular valores na tela de critérios
    def simular_valores_criterios_subcriterios(self, criterios, subcriterios_default, num_frameworks, num_avaliadores):
        print("Simulação ativada")
        for criterio in criterios:
            for entry_peso in self.pesos_criterios[criterio]:
                entry_peso.delete(0, tk.END)
                entry_peso.insert(0, str(random.randint(1, 5)))  # Simula peso de 1 a 5

            sub_pesos = self.pesos_subcriterios[criterio]
            sub_pontuacoes = self.pontuacoes_subcriterios[criterio]
            
            for subcriterio in sub_pesos.keys():
                for entry_peso in sub_pesos[subcriterio]:
                    entry_peso.delete(0, tk.END)
                    entry_peso.insert(0, str(random.randint(1, 5)))  # Simula peso dos subcritérios de 1 a 5

                for pontuacao in sub_pontuacoes[subcriterio]:
                    pontuacao.set(random.randint(1, 5))  # Simula pontuação de 1 a 5

        
    def voltar_para_identificacao(self):
        self.criterios_frame.destroy()  # Destruir a tela de critérios
        self.criar_tela_identificacao()  # Voltar para a tela de identificação


    def voltar_para_identificacao(self):
        self.criterios_frame.destroy()  # Destruir a tela de critérios
        criar_tela_identificacao(self)  # Voltar para a tela de identificação

    def proxima_tela_criterios(self, num_frameworks):
        try:
            # Certificar-se de que todos os nomes de avaliadores e frameworks estão preenchidos
            self.nomes_avaliadores = [entry.get() for entry in self.nomes_avaliadores_entries]
            self.nomes_frameworks = [entry.get() for entry in self.nomes_frameworks_entries]

            if len(self.nomes_avaliadores) < self.num_avaliadores_val or len(self.nomes_frameworks) < self.num_frameworks_val:
                raise ValueError("Por favor, preencha o nome de todos os avaliadores e frameworks.")

            # Destruir a tela anterior (nomes_frame)
            if hasattr(self, 'nomes_frame'):
                self.nomes_frame.destroy()

            # Criar a nova tela de avaliação
            self.iniciar_avaliacao(0)

        except ValueError as e:
            messagebox.showerror("Erro", str(e))
        except Exception as e:
            messagebox.showerror("Erro inesperado", str(e))



    def calcular_notas(self):
        try:
            # Verifica se os avaliadores e frameworks estão definidos corretamente
            if not hasattr(self, 'nomes_avaliadores') or not hasattr(self, 'nomes_frameworks'):
                raise ValueError("Os avaliadores e frameworks não foram definidos corretamente.")

            nomes_avaliadores = self.nomes_avaliadores
            nomes_frameworks = self.nomes_frameworks

            if not nomes_avaliadores or not nomes_frameworks:
                raise ValueError("Os avaliadores ou frameworks não foram preenchidos corretamente.")

            frameworks = []  # Lista para armazenar os dados dos frameworks

            # Iterar sobre todos os frameworks
            for fw_idx in range(self.num_frameworks_val):
                avaliacoes_framework = []  # Armazena as avaliações de cada framework

                # Iterar sobre todos os avaliadores para o framework atual
                for avaliador_idx, avaliador in enumerate(nomes_avaliadores):
                    print(f"Processando: Framework {fw_idx + 1}, Avaliador: {avaliador}")

                    try:
                        # Coletar pesos e pontuações de cada avaliador e framework
                        pesos_criterios = {criterio: int(self.pesos_criterios[criterio][avaliador_idx].get()) for criterio in self.criterios}
                        pontuacoes_sub = {
                            criterio: {subcriterio: int(self.pontuacoes_subcriterios[criterio][subcriterio][avaliador_idx].get())
                                    for subcriterio in self.subcriterios_default[criterio]}
                            for criterio in self.criterios
                        }
                        pesos_sub = {
                            criterio: {subcriterio: int(self.pesos_subcriterios[criterio][subcriterio][avaliador_idx].get())
                                    for subcriterio in self.subcriterios_default[criterio]}
                            for criterio in self.criterios
                        }

                    except Exception as e:
                        raise ValueError(f"Erro ao processar pesos e pontuações para {avaliador}: {e}")

                    try:
                        # Calcula as notas dos critérios para o framework atual
                        notas_criterios = self.calcular_nota_criterio(pesos_sub, pontuacoes_sub)
                    except Exception as e:
                        raise ValueError(f"Erro ao calcular notas para {avaliador}: {e}")

                    try:
                        if len(notas_criterios) == 0:
                            raise ValueError(f"Nenhum critério disponível para {avaliador}")

                        # Criação do DataFrame dos critérios do avaliador
                        df_criterios = pd.DataFrame(list(notas_criterios.items()), columns=["Critério", "Nota do Critério"])
                        df_criterios["Peso do Critério"] = df_criterios["Critério"].apply(lambda x: pesos_criterios[x])

                    except Exception as e:
                        raise ValueError(f"Erro ao gerar DataFrame df_criterios para {avaliador}: {e}")

                    # Coleta os detalhes dos subcritérios e suas notas
                    detalhes_subcriterios = [
                        (criterio, subcriterio, pontuacoes_sub[criterio][subcriterio], pesos_sub[criterio][subcriterio])
                        for criterio in self.criterios
                        for subcriterio in self.subcriterios_default[criterio]
                    ]
                    df_subcriterios = pd.DataFrame(detalhes_subcriterios, columns=["Critério", "Subcriterio", "Pontuação Atribuída", "Peso do Subcriterio"])

                    # Calcula a nota final do avaliador para o framework atual
                    nota_final_avaliador = self.calcular_nota_final(pesos_criterios, notas_criterios)

                    # Adiciona as avaliações do avaliador ao framework
                    avaliacoes_framework.append({
                        'Avaliador': avaliador,
                        'df_criterios': df_criterios,
                        'df_subcriterios': df_subcriterios,
                        'Nota Final': nota_final_avaliador  # Certifique-se de que a nota final é recalculada para cada framework
                    })

                # Adiciona o framework e suas avaliações à lista final de frameworks
                frameworks.append({
                    'Framework': f"Framework {fw_idx + 1}",
                    'avaliacoes': avaliacoes_framework
                })

            # Após coletar todos os dados, gerar o relatório
            self.gerar_relatorio_pdf_e_xlsx(frameworks)

        except ValueError as e:
            messagebox.showerror("Erro", str(e))
        except Exception as e:
            messagebox.showerror("Erro inesperado", str(e))



    
    # Função para simular valores de pesos e pontuações
    def simular_valores(pesos_criterios, pesos_subcriterios, pontuacoes_subcriterios, criterios, subcriterios_default):
        for criterio in criterios:
            # Simular valores de peso dos critérios
            for entry_peso in pesos_criterios[criterio]:
                entry_peso.delete(0, tk.END)  # Limpar o campo
                entry_peso.insert(0, str(random.randint(1, 5)))  # Inserir valor aleatório de 1 a 5

            sub_pesos = pesos_subcriterios[criterio]
            sub_pontuacoes = pontuacoes_subcriterios[criterio]
            
            # Simular valores de peso e pontuação dos subcritérios
            for subcriterio in sub_pesos.keys():
                # Simular pesos dos subcritérios
                for entry_peso in sub_pesos[subcriterio]:
                    entry_peso.delete(0, tk.END)  # Limpar o campo
                    entry_peso.insert(0, str(random.randint(1, 5)))  # Inserir valor aleatório de 1 a 5
                
                # Simular pontuação dos subcritérios
                for pontuacao_var in sub_pontuacoes[subcriterio]:
                    pontuacao_var.set(random.randint(1, 5))  # Definir valor aleatório de 1 a 5

    # Função para calcular as notas de cada critério com base nos pesos e pontuações dos subcritérios
    def calcular_nota_criterio(self, pesos_subcriterios, pontuacoes_subcriterios):
        notas_criterios = {}
        for criterio, subcriterios in pesos_subcriterios.items():
            soma_ponderada = 0
            soma_pesos = 0
            for subcriterio, peso in subcriterios.items():
                # Recupera pontuação ou usa o valor se for IntVar
                pontuacao = pontuacoes_subcriterios[criterio][subcriterio].get() if isinstance(pontuacoes_subcriterios[criterio][subcriterio], IntVar) else pontuacoes_subcriterios[criterio][subcriterio]
                soma_ponderada += pontuacao * peso
                soma_pesos += peso
            if soma_pesos > 0:
                notas_criterios[criterio] = soma_ponderada / soma_pesos
            else:
                notas_criterios[criterio] = 0
        return notas_criterios

    # Função para calcular a nota final com base nos critérios
    def calcular_nota_final(self, pesos_criterios, notas_criterios):
        soma_ponderada = sum(notas_criterios[criterio] * peso for criterio, peso in pesos_criterios.items())
        soma_pesos = sum(pesos_criterios.values())
        nota_final = soma_ponderada / soma_pesos if soma_pesos > 0 else 0
        return nota_final

    def gerar_relatorio_pdf_e_xlsx(self, frameworks):
        try:
            # Solicitar o nome do documento ao usuário
            nome_documento = simpledialog.askstring("Nome do Documento", "Como você gostaria de nomear o documento?")
            if not nome_documento:
                messagebox.showerror("Erro", "Você deve fornecer um nome para o documento.")
                return

            # Criar a pasta 'relatorios' se não existir
            pasta_relatorios = "relatorios"
            if not os.path.exists(pasta_relatorios):
                os.makedirs(pasta_relatorios)

            # Definir caminhos dos arquivos
            filename_pdf = os.path.join(pasta_relatorios, f"{nome_documento}.pdf")
            filename_xlsx = os.path.join(pasta_relatorios, f"{nome_documento}.xlsx")

            # Geração do PDF
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('Arial', 'B', 16)
            pdf.cell(200, 10, 'Relatório de Avaliação', ln=True, align='C')
            pdf.ln(10)

            # Inicializar dados para o PDF
            for framework in frameworks:
                pdf.set_font('Arial', 'B', 14)
                pdf.cell(0, 10, f"Framework: {framework['Framework']}", ln=True)
                pdf.ln(5)

                for avaliacao in framework['avaliacoes']:
                    avaliador = avaliacao['Avaliador']
                    df_criterios = avaliacao['df_criterios']
                    df_subcriterios = avaliacao['df_subcriterios']
                    nota_final = avaliacao['Nota Final']

                    # Título do avaliador
                    pdf.set_font('Arial', 'B', 12)
                    pdf.cell(0, 10, f"Avaliador: {avaliador}", ln=True)
                    pdf.ln(5)

                    # Tabela de critérios
                    pdf.set_font('Arial', 'B', 10)
                    pdf.cell(40, 10, "Critério", 1)
                    pdf.cell(40, 10, "Nota", 1)
                    pdf.cell(40, 10, "Peso", 1)
                    pdf.ln()

                    # Preencher as notas para cada critério
                    for i in range(len(df_criterios)):
                        criterio = df_criterios.iloc[i]['Critério']
                        nota = df_criterios.iloc[i]['Nota do Critério']
                        peso = df_criterios.iloc[i]['Peso do Critério']
                        pdf.set_font('Arial', '', 10)
                        pdf.cell(40, 10, criterio, 1)
                        pdf.cell(40, 10, str(nota), 1)
                        pdf.cell(40, 10, str(peso), 1)
                        pdf.ln()

                        # Adicionar notas dos subcritérios
                        pdf.set_font('Arial', 'B', 10)
                        pdf.cell(40, 10, "Subcritério", 1)
                        pdf.cell(40, 10, "Pontuação", 1)
                        pdf.cell(40, 10, "Peso", 1)
                        pdf.ln()

                        # Filtrar os subcritérios para o critério atual
                        sub_df = df_subcriterios[df_subcriterios['Critério'] == criterio]
                        for j in range(len(sub_df)):
                            subcriterio = sub_df.iloc[j]['Subcriterio']
                            pontuacao = sub_df.iloc[j]['Pontuação Atribuída']
                            peso_sub = sub_df.iloc[j]['Peso do Subcriterio']
                            pdf.set_font('Arial', '', 10)
                            pdf.cell(40, 10, subcriterio, 1)
                            pdf.cell(40, 10, str(pontuacao), 1)
                            pdf.cell(40, 10, str(peso_sub), 1)
                            pdf.ln()

                    pdf.ln(5)
                    # Nota final do avaliador
                    pdf.set_font('Arial', 'B', 12)
                    pdf.cell(0, 10, f"Nota Final do Avaliador {avaliador}: {nota_final:.2f}", ln=True)
                    pdf.ln(10)

            pdf.output(filename_pdf)

            # Geração do XLSX
            with pd.ExcelWriter(filename_xlsx, engine='openpyxl') as writer:
                # Estilos no Excel
                def aplicar_estilos_excel(writer, df, sheet_name):
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    workbook = writer.book
                    sheet = workbook[sheet_name]
                    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    bold_font = Font(bold=True)
                    center_alignment = Alignment(horizontal="center")
                    for row in sheet.iter_rows(min_row=1, max_row=1):
                        for cell in row:
                            cell.fill = fill
                            cell.font = bold_font
                            cell.alignment = center_alignment
                    sheet.freeze_panes = sheet['A2']  # Congelar primeira linha

                # Criar a planilha para cada framework e avaliador
                for framework in frameworks:
                    for avaliacao in framework['avaliacoes']:
                        avaliador = avaliacao['Avaliador']
                        df_criterios = avaliacao['df_criterios']
                        df_criterios.to_excel(writer, sheet_name=f'{framework["Framework"]}_{avaliador}_Criterios', index=False)
                        aplicar_estilos_excel(writer, df_criterios, f'{framework["Framework"]}_{avaliador}_Criterios')

                        df_subcriterios = avaliacao['df_subcriterios']
                        df_subcriterios.to_excel(writer, sheet_name=f'{framework["Framework"]}_{avaliador}_Subcriterios', index=False)
                        aplicar_estilos_excel(writer, df_subcriterios, f'{framework["Framework"]}_{avaliador}_Subcriterios')

                # Criar a planilha comparativa final
                comparativo_data = []
                for framework in frameworks:
                    for avaliacao in framework['avaliacoes']:
                        comparativo_data.append([
                            framework['Framework'], 
                            avaliacao['Avaliador'], 
                            avaliacao['Nota Final']
                        ])

                df_comparativo = pd.DataFrame(comparativo_data, columns=["Framework", "Avaliador", "Nota Final"])
                aplicar_estilos_excel(writer, df_comparativo, 'Comparativo Final')

            messagebox.showinfo("Relatórios Gerados", f"Relatórios PDF e Excel gerados com sucesso:\n{filename_pdf}\n{filename_xlsx}")

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar os relatórios: {e}")




if __name__ == "__main__":
    app = App()
    app.mainloop()
